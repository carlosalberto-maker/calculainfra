"""
Calculadora de infraestructura para EDS UM.

Lee un archivo Excel (plantilla "Información implementación EDS UM.xlsx") con
las pestañas Servicios, Personal (cuerpo de gobierno) y Red, y reporta por
cada servicio y global cuántos equipos de cómputo, impresoras, access points
y switches están instalados, se requieren y faltan.

CRITERIOS OFICIALES IMSS-Bienestar (codificados, están como imagen en Excel):
  Urgencias: triage/mod admisión=1, consultorio=1, cama/5, central enf=1
  Hospitalización: cama/5, central enf=1, farmacia hosp=1
  Quirófano: sala/2, central enf=1
  Tococirugía: sala exp/2, cama/5, central enf=1
  Consulta externa: consultorio=1, especialidad=1, farmacia/ventanilla=1
  Cuerpo de gobierno: 1 por perfil con personal
  Impresoras: 1 cada 10 equipos

CRITERIOS TÉCNICOS (no oficiales, buenas prácticas):
  Access Point: 1 cada 25 dispositivos
  Switch 24p: 1 cada 22 nodos (24 puertos - 2 uplink)

La hoja Red tiene DOS bloques:
  Izquierdo (B-D): AP actual | Switch actual | Obs
  Derecho (E-G):   AP faltante | Switch faltante | Obs → el script escribe aquí
"""

import argparse
import math
import os
import sys
import unicodedata
from datetime import date

import pandas as pd


# ---------------------------------------------------------------------------
# Parámetros de cálculo (ajustables)
# ---------------------------------------------------------------------------
IMPRESORAS_POR_EQUIPOS = 10   # 1 impresora por cada 10 equipos de cómputo
COMPUTADORAS_POR_AP    = 25   # 1 AP por cada 25 dispositivos
PUERTOS_POR_SWITCH     = 24   # switches de 24 puertos
PUERTOS_DE_RESERVA     = 2    # puertos reservados para uplink / crecimiento

# Ya no se usa whitelist fija: se asignan computadoras a todos los perfiles
# del personal que tengan al menos 1 persona (criterio oficial).
# Se conserva como fallback por si no se encuentra la hoja Personal.
PERFILES_CUERPO_GOBIERNO = [
    "direccion", "subdireccion", "jefatura", "epidemiologia",
    "capturista", "archivista", "operador", "farmacia",
    "triage", "estomatologia", "maxilofacial",
]


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def normalizar(texto) -> str:
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = texto.encode("ascii", errors="ignore").decode("ascii")
    return " ".join(texto.split())


def a_numero(valor) -> float:
    try:
        n = float(valor)
        return 0.0 if math.isnan(n) else n
    except (TypeError, ValueError):
        return 0.0


def _es_numero_valido(texto) -> bool:
    try:
        float(texto)
        return True
    except (ValueError, TypeError):
        return False


def clave_servicio(nombre) -> str:
    return normalizar(nombre).replace(" ", "")


def _encontrar_header_row(xls, sheet_name, required_groups):
    df_raw = pd.read_excel(xls, sheet_name, header=None)
    for i, row in df_raw.iterrows():
        row_text = " ".join([normalizar(c) for c in row if pd.notna(c)])
        if not row_text:
            continue
        if all(any(normalizar(v) in row_text for v in group)
               for group in required_groups):
            return i
    return 0


def _encontrar_hoja(xls, *variantes):
    names_lower = {s.lower(): s for s in xls.sheet_names}
    for v in variantes:
        if v in xls.sheet_names:
            return v
    for v in variantes:
        if v.lower() in names_lower:
            return names_lower[v.lower()]
    for v in variantes:
        vl = v.lower()
        for sl, sr in names_lower.items():
            if vl in sl:
                return sr
    return variantes[0]


def _encontrar_columna(df, *variantes):
    cols_norm = {c: normalizar(c) for c in df.columns}
    for c, c_norm in cols_norm.items():
        for v in variantes:
            if normalizar(v) in c_norm:
                return c
    return None


# ---------------------------------------------------------------------------
# Criterios: equipos de cómputo requeridos por (servicio, área, cantidad)
# ---------------------------------------------------------------------------
def calcular_computadoras_area(servicio, area, cantidad) -> int:
    """
    Aplica los criterios oficiales de la pestaña Criterios_estimación.

    CRITERIOS IMPLEMENTADOS:
    URGENCIAS / ADMISIÓN CONTINUA
      Módulo admisión / triage            1 x módulo
      Consultorios primer contacto        1 x consultorio
      Camas (observación, choque, etc.)   1 x 5 camas
      Central de enfermería               1 x central
      Jefatura servicio / enfermería      1 fijo

    HOSPITALIZACIÓN (incluye terapias intensivas)
      Camas hosp., UCIA, UCIP, UCIN,
        cirugía ambulatoria               1 x 5 camas
      Cunero patológico                   1 x 5 cunas
      Recuperación post-parto/quirúrg.    1 x central
      Central de enfermería               1 x central
      Farmacia hospitalaria               1 x área
      Jefatura servicio / enfermería      1 fijo

    QUIRÓFANO
      Salas de operación                  1 x 2 salas
      Central de enfermería               1 x central
      Jefatura servicio / enfermería      1 fijo

    TOCOCIRUGÍA
      Salas de operación/expulsión        1 x 2 salas
      Camas labor, recuperación           1 x 5 camas
      Central de enfermería               1 x central
      Jefatura servicio / enfermería      1 fijo

    CONSULTA EXTERNA
      Consultorios (por especialidad)     1 x consultorio
      Farmacia hospitalaria               1 x área
      Farmacia CE / ventanilla            1 x ventanilla
      Control enfermería                  1 x área
      Auxiliar administrativo             1 x 2 salas

    ESTANCIA CORTA (hemodiálisis, quimioterapia…)
      Camas / sillones                    1 x 5 camas
      Central de enfermería               1 x central
      Jefatura servicio / enfermería      1 fijo
    """
    serv = normalizar(servicio)
    ar   = normalizar(area)
    cant = a_numero(cantidad)

    if cant <= 0:
        return 0

    # Filas "Áreas de …" son agrupadores sin equipos propios
    # solo si cantidad es 0 o no tienen keywords que indiquen área funcional
    if ar.startswith("area de ") or ar.startswith("areas de "):
        if cant <= 0:
            return 0
        # Si tiene cantidad >0, se deja pasar a las reglas por servicio

    # ── Urgencias / admisión continua ──────────────────────────────────
    if "urgencia" in serv or "admision continua" in serv:
        if "triage" in ar or "modulo de admision" in ar or ar.startswith("admision"):
            return math.ceil(cant)
        if "consultorio" in ar:
            return math.ceil(cant)
        if any(k in ar for k in ("cama", "cunero", "cuna", "choque", "observacion", "reanimacion")):
            return math.ceil(cant / 5)
        if "central de enfermeria" in ar or "centrales de enfermeria" in ar:
            return math.ceil(cant)
        if "jefatura" in ar:
            return 1
        return 0

    # ── Hospitalización (incluye terapias intensivas e intermedias) ────
    if "hospitalizacion" in serv:
        if any(k in ar for k in ("cama", "cunero", "cuna", "recuperacion",
                                  "ucia", "ucip", "ucin", "cirugia ambulatoria")):
            return math.ceil(cant / 5)
        if "farmacia hospitalaria" in ar:
            return math.ceil(cant)
        if "central de enfermeria" in ar or "recuperacion post" in ar:
            return math.ceil(cant)
        if "jefatura" in ar:
            return 1
        return 0

    # ── Quirófano ──────────────────────────────────────────────────────
    if "quirofano" in serv:
        if "central de enfermeria" in ar:
            return math.ceil(cant)
        if "jefatura" in ar:
            return 1
        if any(k in ar for k in ("sala", "quirofano")):
            return math.ceil(cant / 2)
        if "cama" in ar or "recuperacion" in ar:
            return math.ceil(cant / 5)
        return 0

    # ── Tococirugía / Toco cirugía ─────────────────────────────────────
    if "tococirugia" in serv or "toco cirugia" in serv:
        if "central de enfermeria" in ar:
            return math.ceil(cant)
        if "jefatura" in ar:
            return 1
        if any(k in ar for k in ("cama", "recuperacion", "labor de parto")):
            return math.ceil(cant / 5)
        if any(k in ar for k in ("sala", "expulsion", "tococirug")):
            return math.ceil(cant / 2)
        return 0

    # ── Consulta externa ───────────────────────────────────────────────
    if "consulta externa" in serv:
        if "consultorio" in ar:
            return math.ceil(cant)
        if "farmacia hospitalaria" in ar:
            return math.ceil(cant)
        if "ventanilla" in ar and "farmacia" in ar:
            return math.ceil(cant)
        if "ventanilla" in ar and "archivo" in ar:
            return math.ceil(cant)
        if "control" in ar and "enfermeria" in ar:
            return math.ceil(cant)
        if "auxiliar" in ar or "administrativo" in ar:
            return math.ceil(cant / 2)
        ESPECIALIDADES_CE = (
            "ginecologia", "medicina interna", "cirugia", "pediatria",
            "psicologia", "odontologia", "nutricion", "oftalmologia",
            "traumatologia", "ortopedia", "dermatologia", "cardiologia",
            "neurologia", "urologia", "otorrinolaringologia", "consulta general",
            "medicina familiar", "geriatria", "oncologia", "endocrinologia",
            "reumatologia", "nefrologia", "neumologia", "hematologia",
            "infectologia", "gastroenterologia",
        )
        if any(esp in ar for esp in ESPECIALIDADES_CE):
            return math.ceil(cant)
        return 0

    # ── Estancia corta (hemodiálisis, quimioterapia…) ─────────────────
    if any(k in serv for k in ("estancia corta", "hemodialisis", "quimioterapia")):
        if any(k in ar for k in ("cama", "sillon", "puesto")):
            return math.ceil(cant / 5)
        if "central de enfermeria" in ar:
            return math.ceil(cant)
        if "jefatura" in ar:
            return 1
        return 0

    # ── Fallback genérico por tipo de área ────────────────────────────
    if "cama" in ar or "cuna" in ar:
        return math.ceil(cant / 5)
    if "consultorio" in ar:
        return math.ceil(cant)
    if "central de enfermeria" in ar:
        return math.ceil(cant)
    if "jefatura" in ar:
        return 1
    if "modulo" in ar:
        return math.ceil(cant)

    return 0


# ---------------------------------------------------------------------------
# Lectura del Excel
# ---------------------------------------------------------------------------
def parse_excel(file_path):
    print(f"Cargando archivo: {file_path}")
    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")
        sys.exit(1)

    def leer_hoja(nombre, expected_terms):
        header_row = _encontrar_header_row(xls, nombre, expected_terms)
        if header_row > 0:
            print(f"  >> Encabezados detectados en fila {header_row + 1}")
        try:
            df = pd.read_excel(xls, nombre, header=header_row)
            df.columns = [str(c).strip() for c in df.columns]
            return df, header_row
        except Exception as e:
            print(f"  ! No se pudo leer '{nombre}': {e}")
            return pd.DataFrame(), 0

    print("--- Leyendo Servicios ---")
    servicios, hr_serv = leer_hoja("Servicios", [["Servicio"], ["Area"], ["Cantidad"]])
    if not servicios.empty and "Servicio" in servicios.columns:
        servicios["Servicio"] = servicios["Servicio"].ffill()
        subset = [c for c in ("Servicio", "Área", "Area") if c in servicios.columns]
        if subset:
            servicios = servicios.dropna(subset=subset, how="all")

    personal_sheet = _encontrar_hoja(xls, "Personal (cuerpo de gobierno)", "Personal", "personal")
    print(f"--- Leyendo {personal_sheet} ---")
    personal, hr_pers = leer_hoja(personal_sheet, [["Perfil"], ["Turno"], ["personal"]])
    if not personal.empty:
        cols_perfil = [c for c in personal.columns if "Perfil" in c]
        if cols_perfil:
            personal[cols_perfil[0]] = personal[cols_perfil[0]].ffill()
            personal = personal.dropna(subset=[cols_perfil[0]], how="all")

    print("--- Leyendo Red ---")
    red, hr_red = leer_hoja("Red", [["Servicio"], ["Access", "Acces"], ["Switch"]])
    if not red.empty and "Servicio" in red.columns:
        red["Servicio"] = red["Servicio"].ffill()
        red = red.dropna(subset=["Servicio"], how="all")

    header_rows = {"Servicios": hr_serv, "Personal": hr_pers, "Red": hr_red}
    return servicios, personal, red, header_rows


# ---------------------------------------------------------------------------
# Identificar columnas de la hoja Red (bloque izquierdo vs derecho)
# ---------------------------------------------------------------------------
def _identificar_columnas_red(red_df):
    """
    La hoja Red tiene DOS bloques con nombres iguales.
    pandas renombra la segunda aparición con sufijo '.1'
    Bloque izquierdo (actual):  AP, Switch, Observaciones      → sin sufijo
    Bloque derecho  (faltante): AP.1, Switch.1, Observaciones.1 → con '.1'
    """
    col_ap_actual  = col_sw_actual  = col_obs_actual  = None
    col_ap_faltante = col_sw_faltante = col_obs_faltante = None

    for c in red_df.columns:
        cl = str(c).lower()
        tiene_sufijo = str(c).endswith(".1")
        if "acces" in cl or "access" in cl:
            if not tiene_sufijo and col_ap_actual is None:
                col_ap_actual = c
            elif tiene_sufijo and col_ap_faltante is None:
                col_ap_faltante = c
        if "switch" in cl:
            if not tiene_sufijo and col_sw_actual is None:
                col_sw_actual = c
            elif tiene_sufijo and col_sw_faltante is None:
                col_sw_faltante = c
        if "observaciones" in cl or "obs" in cl:
            if not tiene_sufijo and col_obs_actual is None:
                col_obs_actual = c
            elif tiene_sufijo and col_obs_faltante is None:
                col_obs_faltante = c

    return (col_ap_actual, col_sw_actual, col_obs_actual,
            col_ap_faltante, col_sw_faltante, col_obs_faltante)


# ---------------------------------------------------------------------------
# Estructura de un servicio
# ---------------------------------------------------------------------------
def _servicio_nuevo(nombre):
    return {
        "nombre":        str(nombre).strip(),
        "comp_actual":   0,
        "comp_req":      0,
        "imp_actual":    0,
        "ap_actual":     0,
        "ap_actual_raw": "",
        "sw_actual":     0,
        "sw_actual_raw": "",
        "obs_actual":    "",
        "_detalle":      [],
    }


def _agregar_detalle(s, tipo, fuente, valor, explicacion=""):
    s["_detalle"].append({
        "tipo":        tipo,
        "fuente":      str(fuente),
        "valor":       valor,
        "explicacion": explicacion,
    })


# ---------------------------------------------------------------------------
# Cálculo principal
# ---------------------------------------------------------------------------
def calcular(servicios_df, personal_df, red_df):
    por_servicio = {}

    def get(nombre):
        clave = clave_servicio(nombre)
        if not clave:
            return None
        if clave not in por_servicio:
            por_servicio[clave] = _servicio_nuevo(nombre)
        else:
            existente = por_servicio[clave]["nombre"]
            nuevo = str(nombre).strip()
            if len(nuevo) > len(existente):
                por_servicio[clave]["nombre"] = nuevo
        return por_servicio[clave]

    # ── 1. Hoja Servicios → comp_req, comp_actual, imp_actual ──────────
    if not servicios_df.empty:
        col_comp_actual = _encontrar_columna(
            servicios_df,
            "Número de equipos de cómputo (actual)",
            "Número de equipos de cómputo (actialmente)",
            "equipos de cómputo",
            "# equipos computo",
        )
        col_imp_actual = _encontrar_columna(
            servicios_df, "Número de impresoras", "# impresoras")
        col_area = _encontrar_columna(servicios_df, "Área", "Area") or "Área"
        col_cant = _encontrar_columna(servicios_df, "Cantidad", "número de") or "Cantidad (número de)"

        for _, row in servicios_df.iterrows():
            serv_name = row.get("Servicio", "")
            s = get(serv_name)
            if s is None:
                continue
            area = row.get(col_area, "")
            cant = row.get(col_cant, 0)

            req = calcular_computadoras_area(serv_name, area, cant)
            s["comp_req"] += req
            if req:
                _agregar_detalle(s, "comp_req",
                                 f"{serv_name} — {area}",
                                 req,
                                 f"criterio aplicado sobre cantidad={int(a_numero(cant))}")

            if col_comp_actual:
                val = int(a_numero(row.get(col_comp_actual, 0)))
                s["comp_actual"] += val
                if val:
                    _agregar_detalle(s, "comp_actual",
                                     f"{serv_name} — {area}",
                                     val,
                                     "reportado en hoja Servicios")

            if col_imp_actual:
                val = int(a_numero(row.get(col_imp_actual, 0)))
                s["imp_actual"] += val
                if val:
                    _agregar_detalle(s, "imp_actual",
                                     f"{serv_name} — {area}",
                                     val,
                                     "reportado en hoja Servicios")

    # ── 2. Hoja Personal → servicio virtual "Cuerpo de gobierno" ──────
    if not personal_df.empty:
        cg = get("Cuerpo de gobierno")
        cols_perfil   = [c for c in personal_df.columns if "Perfil" in c]
        col_pers_cant = _encontrar_columna(personal_df, "Cantidad de personal", "personal")
        col_comp_pers = _encontrar_columna(personal_df,
                                            "# equipos computo por perfil",
                                            "# equipos computo totales",
                                            "equipos computo")
        col_imp_pers  = _encontrar_columna(personal_df,
                                            "# impresoras por perfil",
                                            "# impresoras totales",
                                            "impresoras")

        if col_comp_pers:
            val = int(pd.to_numeric(personal_df[col_comp_pers],
                                    errors="coerce").fillna(0).sum())
            cg["comp_actual"] += val
            if val:
                _agregar_detalle(cg, "comp_actual",
                                 "Personal (cuerpo de gobierno)",
                                 val,
                                 f"suma de columna '{col_comp_pers}'")
        if col_imp_pers:
            val = int(pd.to_numeric(personal_df[col_imp_pers],
                                    errors="coerce").fillna(0).sum())
            cg["imp_actual"] += val
            if val:
                _agregar_detalle(cg, "imp_actual",
                                 "Personal (cuerpo de gobierno)",
                                 val,
                                 f"suma de columna '{col_imp_pers}'")

        if cols_perfil and col_pers_cant:
            col_perfil = cols_perfil[0]
            personal_df = personal_df.copy()
            personal_df[col_pers_cant] = pd.to_numeric(
                personal_df[col_pers_cant], errors="coerce").fillna(0)
            agrupado = personal_df.groupby(col_perfil)[col_pers_cant].sum()
            for perfil, cant_personal in agrupado.items():
                if cant_personal <= 0:
                    continue
                # Criterio oficial: 1 equipo por cada perfil o puesto
                # de jefatura y personal operativo con al menos 1 persona
                cg["comp_req"] += 1
                _agregar_detalle(cg, "comp_req",
                                 f"Personal — {perfil}",
                                 1,
                                 f"perfil con {int(cant_personal)} personas en total")

    # ── 3. Hoja Red → ap_actual, sw_actual ────────────────────────────
    if not red_df.empty:
        (col_ap_act, col_sw_act, col_obs_act,
         col_ap_falt, col_sw_falt, col_obs_falt) = _identificar_columnas_red(red_df)

        for _, row in red_df.iterrows():
            serv_name = row.get("Servicio", "")
            s = get(serv_name)
            if s is None:
                continue
            if col_ap_act:
                raw = row.get(col_ap_act)
                raw = "" if pd.isna(raw) else str(raw).strip()
                s["ap_actual_raw"] = raw
                val = int(a_numero(raw))
                s["ap_actual"] += val
                if val:
                    _agregar_detalle(s, "ap_actual",
                                     f"{serv_name} (Red)", val,
                                     "AP actual reportado en hoja Red")
            if col_sw_act:
                raw = row.get(col_sw_act)
                raw = "" if pd.isna(raw) else str(raw).strip()
                s["sw_actual_raw"] = raw
                val = int(a_numero(raw))
                s["sw_actual"] += val
                if val:
                    _agregar_detalle(s, "sw_actual",
                                     f"{serv_name} (Red)", val,
                                     "Switch actual reportado en hoja Red")
            if col_obs_act:
                raw = row.get(col_obs_act)
                s["obs_actual"] = "" if pd.isna(raw) else str(raw).strip()

    # ── 4. Calcular requeridos (AP, SW por servicio; IMP global) ──────
    puertos_utiles = max(PUERTOS_POR_SWITCH - PUERTOS_DE_RESERVA, 1)
    for s in por_servicio.values():
        s["imp_req"] = 0  # se calcula globalmente abajo
        s["ap_req"]  = math.ceil(s["comp_req"] / COMPUTADORAS_POR_AP)    if s["comp_req"] else 0
        nodos        = s["comp_req"] + s["ap_req"]
        s["sw_req"]  = math.ceil(nodos / puertos_utiles) if nodos else 0

    # Impresoras: criterio global 1:10 sobre el total de computadoras
    total_comp = sum(s["comp_req"] for s in por_servicio.values())
    global_imp = math.ceil(total_comp / IMPRESORAS_POR_EQUIPOS) if total_comp else 0
    # Distribuir proporcionalmente
    servicios_imp = [s for s in por_servicio.values() if s["comp_req"] > 0]
    if servicios_imp:
        # Asignar según proporción de comp_req
        comp_sum = sum(s["comp_req"] for s in servicios_imp)
        shares = []
        for s in servicios_imp:
            exact = s["comp_req"] * global_imp / comp_sum
            base = math.floor(exact)
            shares.append((exact - base, base))
        asignado = sum(b for _, b in shares)
        resto = global_imp - asignado
        indices = sorted(range(len(servicios_imp)), key=lambda i: shares[i][0], reverse=True)
        for idx in indices[:resto]:
            shares[idx] = (shares[idx][0], shares[idx][1] + 1)
        for i, s in enumerate(servicios_imp):
            s["imp_req"] = shares[i][1]

    # ── 5. Totales globales ────────────────────────────────────────────
    claves = ("Computadoras", "Impresoras", "Access Point", "Switch")
    actual    = {k: 0 for k in claves}
    requerido = {k: 0 for k in claves}
    for s in por_servicio.values():
        actual["Computadoras"]    += s["comp_actual"]
        actual["Impresoras"]      += s["imp_actual"]
        actual["Access Point"]    += s["ap_actual"]
        actual["Switch"]          += s["sw_actual"]
        requerido["Computadoras"] += s["comp_req"]
        requerido["Impresoras"]   += s["imp_req"]
        requerido["Access Point"] += s["ap_req"]
        requerido["Switch"]       += s["sw_req"]

    return actual, requerido, por_servicio


# ---------------------------------------------------------------------------
# Aplicar overrides de totales (ej. valores del Excel resumen oficial)
# ---------------------------------------------------------------------------
def aplicar_overrides(actual, requerido, por_servicio, overrides=None):
    """
    Si el usuario proporciona valores objetivo (overrides), se escalan
    proporcionalmente los valores por servicio y se recalcula todo.
    overrides: dict con claves {comp_req, imp_req, ap_req, sw24_req}
    """
    if not overrides:
        overrides = {}

    campo_a_clave = {
        "comp_req": ("Computadoras", "comp_req"),
        "imp_req":  ("Impresoras",   "imp_req"),
        "ap_req":   ("Access Point", "ap_req"),
        "sw24_req": ("Switch",       "sw_req"),
    }

    comp_was_overridden = "comp_req" in overrides

    for override_key, (clave_global, campo_servicio) in campo_a_clave.items():
        target = overrides.get(override_key)
        if target is None:
            continue
        target = int(target)
        current = sum(s[campo_servicio] for s in por_servicio.values())
        if current == 0 or target == current:
            continue

        # Distribución proporcional (método del resto mayor)
        servicios_lista = list(por_servicio.values())
        n = len(servicios_lista)
        shares = []
        for s in servicios_lista:
            exact = s[campo_servicio] * target / current
            base = math.floor(exact)
            shares.append((exact - base, base))

        asignado = sum(b for _, b in shares)
        resto = target - asignado
        indices = sorted(range(n), key=lambda i: shares[i][0], reverse=True)
        for idx in indices[:resto]:
            shares[idx] = (shares[idx][0], shares[idx][1] + 1)

        for i, s in enumerate(servicios_lista):
            s[campo_servicio] = shares[i][1]

    # Si se sobrescribió comp_req, recalcular imp_req global
    # (criterio oficial: 1 impresora por cada 10 equipos, sobre el total)
    if comp_was_overridden:
        total_comp = sum(s["comp_req"] for s in por_servicio.values())
        global_imp = math.ceil(total_comp / IMPRESORAS_POR_EQUIPOS) if total_comp else 0
        current_imp = sum(s["imp_req"] for s in por_servicio.values())
        if current_imp > 0 and global_imp != current_imp:
            servicios_lista = list(por_servicio.values())
            n = len(servicios_lista)
            shares = []
            for s in servicios_lista:
                exact = s["imp_req"] * global_imp / current_imp if current_imp else 0
                base = math.floor(exact)
                shares.append((exact - base, base))
            asignado = sum(b for _, b in shares)
            resto = global_imp - asignado
            indices = sorted(range(n), key=lambda i: shares[i][0], reverse=True)
            for idx in indices[:resto]:
                shares[idx] = (shares[idx][0], shares[idx][1] + 1)
            for i, s in enumerate(servicios_lista):
                s["imp_req"] = shares[i][1]

    # Si se proporcionó imp_req override, aplicarlo
    if "imp_req" in overrides:
        target_imp = int(overrides["imp_req"])
        current_imp = sum(s["imp_req"] for s in por_servicio.values())
        if current_imp > 0 and target_imp != current_imp:
            servicios_lista = list(por_servicio.values())
            n = len(servicios_lista)
            shares = []
            for s in servicios_lista:
                exact = s["imp_req"] * target_imp / current_imp
                base = math.floor(exact)
                shares.append((exact - base, base))
            asignado = sum(b for _, b in shares)
            resto = target_imp - asignado
            indices = sorted(range(n), key=lambda i: shares[i][0], reverse=True)
            for idx in indices[:resto]:
                shares[idx] = (shares[idx][0], shares[idx][1] + 1)
            for i, s in enumerate(servicios_lista):
                s["imp_req"] = shares[i][1]

    # Recalcular totales tras overrides
    for k in requerido:
        requerido[k] = 0
    for s in por_servicio.values():
        requerido["Computadoras"] += s["comp_req"]
        requerido["Impresoras"]   += s["imp_req"]
        requerido["Access Point"] += s["ap_req"]
        requerido["Switch"]       += s["sw_req"]

    return por_servicio


# ---------------------------------------------------------------------------
# Calcular faltantes, excesos y banderas sin_datos
# ---------------------------------------------------------------------------
def calcular_faltantes_y_excesos(por_servicio, actual, requerido):
    puertos_utiles = max(PUERTOS_POR_SWITCH - PUERTOS_DE_RESERVA, 1)
    for s in por_servicio.values():
        s["comp_falt"] = max(s["comp_req"] - s["comp_actual"], 0)
        s["imp_falt"]  = max(s["imp_req"]  - s["imp_actual"],  0)
        s["ap_falt"]   = max(s["ap_req"]   - s["ap_actual"],   0)
        s["sw_falt"]   = max(s["sw_req"]   - s["sw_actual"],   0)

        s["comp_exceso"] = max(s["comp_actual"] - s["comp_req"], 0)
        s["imp_exceso"]  = max(s["imp_actual"]  - s["imp_req"],  0)
        s["ap_exceso"]   = max(s["ap_actual"]   - s["ap_req"],   0)
        s["sw_exceso"]   = max(s["sw_actual"]   - s["sw_req"],   0)

        sin_datos_servicios = (
            "hospitalizacion", "quirofano", "tococirugia",
            "toco cirugia", "estancia corta"
        )
        serv_norm = normalizar(s["nombre"]).replace(" ", "")
        s["sin_datos"] = (
            s["comp_actual"] == 0 and
            s["comp_req"] == 0 and
            s["imp_actual"] == 0 and
            s["ap_actual"] == 0 and
            s["sw_actual"] == 0 and
            any(k.replace(" ", "") in serv_norm for k in sin_datos_servicios)
        )

    # Recalcular totales
    for k in requerido:
        requerido[k] = 0
    for s in por_servicio.values():
        requerido["Computadoras"] += s["comp_req"]
        requerido["Impresoras"]   += s["imp_req"]
        requerido["Access Point"] += s["ap_req"]
        requerido["Switch"]       += s["sw_req"]

    faltantes = {k: max(requerido[k] - actual[k], 0) for k in requerido}
    excesos   = {k: max(actual[k] - requerido[k], 0) for k in requerido}

    return faltantes, excesos


# ---------------------------------------------------------------------------
# Función conveniencia para Flask / uso como módulo
# ---------------------------------------------------------------------------
def procesar_archivo(file_path, overrides=None):
    """Wrapper: parse_excel + calcular + overrides + faltantes. Devuelve todo."""
    servicios, personal, red, header_rows = parse_excel(file_path)
    actual, requerido, por_servicio = calcular(servicios, personal, red)
    por_servicio = aplicar_overrides(actual, requerido, por_servicio, overrides)
    faltantes, excesos = calcular_faltantes_y_excesos(por_servicio, actual, requerido)
    return actual, requerido, faltantes, excesos, por_servicio, header_rows


# ---------------------------------------------------------------------------
# Reporte en consola
# ---------------------------------------------------------------------------
def generar_reporte(actual, requerido, faltantes, excesos, por_servicio):
    print("\n" + "=" * 72)
    print("REPORTE DE ESTIMACIÓN DE INFRAESTRUCTURA (EDS UM)")
    print("=" * 72)
    print(f"{'Equipo':<18} | {'Actual':>8} | {'Requerido':>10} | {'Faltante':>9} | {'Exceso':>7}")
    print("-" * 72)
    for item in ("Computadoras", "Impresoras", "Access Point", "Switch"):
        exc = excesos.get(item, 0)
        print(f"{item:<18} | {int(actual[item]):>8} | "
              f"{int(requerido[item]):>10} | {int(faltantes[item]):>9} | "
              f"{int(exc):>7}")
    print("=" * 72)

    if por_servicio:
        print("\nDESGLOSE POR SERVICIO  (A=Actual  R=Requerido  F=Faltante  E=Exceso)")
        enc = (f"{'Servicio':<38} | {'Computadoras':^16} | {'Impresoras':^12} | "
               f"{'Access Point':^16} | {'Switch':^16}")
        sub = (f"{'':<38} | {'A':>3} {'R':>3} {'F':>3} {'E':>3} | "
               f"{'A':>2} {'R':>2} {'F':>2} {'E':>2} | "
               f"{'A':>3} {'R':>3} {'F':>3} {'E':>3} | "
               f"{'A':>3} {'R':>3} {'F':>3} {'E':>3}")
        sep = "-" * len(enc)
        print(sep); print(enc); print(sub); print(sep)
        for s in por_servicio.values():
            total_val = s["comp_actual"] + s["comp_req"] + s["imp_actual"] + s["imp_req"] + \
                        s["ap_actual"] + s["ap_req"] + s["sw_actual"] + s["sw_req"]
            if total_val == 0:
                continue  # omitir filas de texto sin datos
            nom = s["nombre"][:35] + "..." if len(s["nombre"]) > 38 else s["nombre"]
            if s.get("sin_datos"):
                print(f"{nom:<38} |  [!] SIN INFORMACION REPORTADA para este servicio")
                continue
            print(
                f"{nom:<38} | "
                f"{s['comp_actual']:>3} {s['comp_req']:>3} {s['comp_falt']:>3} {s['comp_exceso']:>3} | "
                f"{s['imp_actual']:>2} {s['imp_req']:>2} {s['imp_falt']:>2} {s['imp_exceso']:>2} | "
                f"{s['ap_actual']:>3} {s['ap_req']:>3} {s['ap_falt']:>3} {s['ap_exceso']:>3} | "
                f"{s['sw_actual']:>3} {s['sw_req']:>3} {s['sw_falt']:>3} {s['sw_exceso']:>3}"
            )
        print(sep)

    print(f"\nCriterios aplicados:")
    print(f"  · Impresoras : 1 por cada {IMPRESORAS_POR_EQUIPOS} equipos de cómputo")
    print(f"  · Access Point: 1 por cada {COMPUTADORAS_POR_AP} equipos")
    print(f"  · Switch {PUERTOS_POR_SWITCH}p : 1 por cada {PUERTOS_POR_SWITCH - PUERTOS_DE_RESERVA} nodos (2 puertos reservados)")


# ---------------------------------------------------------------------------
# Escritura opcional al bloque derecho de la hoja Red
# ---------------------------------------------------------------------------
def actualizar_excel_red(file_path, por_servicio, header_row=4):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ! openpyxl no instalado.")
        return

    wb = load_workbook(file_path)
    if "Red" not in wb.sheetnames:
        print("  ! La pestaña 'Red' no existe.")
        return
    ws = wb["Red"]

    fila = header_row + 2
    escritos = 0
    sin_fila = []
    while True:
        celda = ws.cell(row=fila, column=1).value
        if celda is None or str(celda).strip() == "":
            break
        clave = clave_servicio(celda)
        if clave in por_servicio:
            s = por_servicio[clave]
            ws.cell(row=fila, column=5, value=s["ap_falt"])
            ws.cell(row=fila, column=6, value=s["sw_falt"])
            obs = (f"AP req={s['ap_req']} (actual={s['ap_actual']}, faltante={s['ap_falt']}); "
                   f"Switch req={s['sw_req']} (actual={s['sw_actual']}, faltante={s['sw_falt']})")
            ws.cell(row=fila, column=7, value=obs)
            escritos += 1
        else:
            sin_fila.append(str(celda).strip())
        fila += 1

    # ── Escribir Cuerpo de gobierno (virtual, sin fila propia en Red) ──
    cg_clave = clave_servicio("Cuerpo de gobierno")
    if cg_clave in por_servicio:
        # Verificar si ya fue escrito como parte de una fila existente
        ya_escrito = False
        for r in range(header_row + 2, fila):
            celda = ws.cell(row=r, column=1).value
            if celda and clave_servicio(celda) == cg_clave:
                ya_escrito = True
                break
        if not ya_escrito:
            s = por_servicio[cg_clave]
            ws.cell(row=fila, column=1, value="Cuerpo de gobierno")
            ws.cell(row=fila, column=5, value=s["ap_falt"])
            ws.cell(row=fila, column=6, value=s["sw_falt"])
            obs_cg = (f"AP req={s['ap_req']} (actual={s['ap_actual']}, faltante={s['ap_falt']}); "
                      f"Switch req={s['sw_req']} (actual={s['sw_actual']}, faltante={s['sw_falt']})")
            ws.cell(row=fila, column=7, value=obs_cg)
            escritos += 1
            print(f"  >> NOTA: 'Cuerpo de gobierno' no existía en hoja Red — se agregó como nueva fila.")

    wb.save(file_path)
    print(f"  >> Excel actualizado: {escritos} servicios escritos en hoja Red (bloque derecho).")
    if sin_fila:
        restantes = [s for s in sin_fila if clave_servicio(s) != cg_clave]
        if restantes:
            print(f"  >> NOTA: {len(restantes)} servicio(s) no tienen fila en hoja Red: {', '.join(restantes)}")


# ---------------------------------------------------------------------------
# Reporte HTML
# ---------------------------------------------------------------------------
def generar_html(file_path, actual, requerido, faltantes, excesos, por_servicio, output_path=None):
    """
    Genera el reporte HTML.
    Siempre devuelve el string HTML.
    Si output_path se proporciona, también guarda el archivo en esa ruta.
    """
    nombre_archivo = os.path.basename(file_path)
    hoy = date.today().strftime("%d/%m/%Y")

    def badge(val, tipo):
        if tipo == "falt" and val > 0:
            return f'<td class="falt">{val} ▲</td>'
        if tipo == "exc" and val > 0:
            return f'<td class="exc">{val} ▼</td>'
        if tipo == "ok":
            return f'<td class="ok">{val}</td>'
        return f'<td>{val}</td>'

    def sin_datos_td():
        return '<td colspan="16" class="sin-datos">⚠ Sin información reportada para este servicio</td>'

    filas_serv = ""
    for s in por_servicio.values():
        total_val = s["comp_actual"] + s["comp_req"] + s["imp_actual"] + s["imp_req"] + \
                    s["ap_actual"] + s["ap_req"] + s["sw_actual"] + s["sw_req"]
        if total_val == 0:
            continue  # omitir filas de texto sin datos

        cls_fila = ' class="row-alert"' if any([
            s["comp_falt"], s["imp_falt"], s["ap_falt"], s["sw_falt"]]) else ""

        obs_partes = []
        if s.get("sin_datos"):
            obs_partes.append("ℹ Sin información reportada — completar hoja Servicios")
        if s["comp_exceso"]:
            obs_partes.append(f"⚠ Exceso de {s['comp_exceso']} computadora(s)")
        if s["obs_actual"]:
            obs_partes.append(s["obs_actual"])
        obs_txt = " | ".join(obs_partes) if obs_partes else "—"

        if s.get("sin_datos"):
            filas_serv += f"""
        <tr class="row-sin-datos">
          <td class="serv-name">{s['nombre']}</td>
          {sin_datos_td()}
        </tr>"""
        else:
            filas_serv += f"""
        <tr{cls_fila}>
          <td class="serv-name">{s['nombre']}</td>
          <td>{s['comp_actual']}</td><td>{s['comp_req']}</td>
          {badge(s['comp_falt'], 'falt')}{badge(s['comp_exceso'], 'exc')}
          <td>{s['imp_actual']}</td><td>{s['imp_req']}</td>
          {badge(s['imp_falt'], 'falt')}{badge(s['imp_exceso'], 'exc')}
          <td>{s['ap_actual']}</td><td>{s['ap_req']}</td>
          {badge(s['ap_falt'], 'falt')}{badge(s['ap_exceso'], 'exc')}
          <td>{s['sw_actual']}</td><td>{s['sw_req']}</td>
          {badge(s['sw_falt'], 'falt')}{badge(s['sw_exceso'], 'exc')}
          <td class="obs">{obs_txt}</td>
        </tr>"""

    def tarjeta(titulo, icono, actual_v, req_v, falt_v, exc_v):
        estado = "card-ok"
        if falt_v > 0:
            estado = "card-falt"
        elif exc_v > 0:
            estado = "card-exc"
        return f"""
        <div class="card {estado}">
          <div class="card-icon">{icono}</div>
          <div class="card-title">{titulo}</div>
          <div class="card-nums">
            <span class="num-actual">{actual_v}<small>actual</small></span>
            <span class="num-req">{req_v}<small>requerido</small></span>
          </div>
          <div class="card-diff">
            {'<span class="badge-falt">▲ Faltan ' + str(falt_v) + '</span>' if falt_v > 0 else ''}
            {'<span class="badge-exc">▼ Exceso ' + str(exc_v) + '</span>' if exc_v > 0 else ''}
            {'<span class="badge-ok">✓ Cubierto</span>' if falt_v == 0 and exc_v == 0 else ''}
          </div>
        </div>"""

    cards = (
        tarjeta("Computadoras", "🖥️",
                int(actual["Computadoras"]), int(requerido["Computadoras"]),
                int(faltantes["Computadoras"]), int(excesos.get("Computadoras", 0))) +
        tarjeta("Impresoras", "🖨️",
                int(actual["Impresoras"]), int(requerido["Impresoras"]),
                int(faltantes["Impresoras"]), int(excesos.get("Impresoras", 0))) +
        tarjeta("Access Points", "📡",
                int(actual["Access Point"]), int(requerido["Access Point"]),
                int(faltantes["Access Point"]), int(excesos.get("Access Point", 0))) +
        tarjeta("Switches (24p)", "🔌",
                int(actual["Switch"]), int(requerido["Switch"]),
                int(faltantes["Switch"]), int(excesos.get("Switch", 0)))
    )

    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Reporte EDS — {nombre_archivo}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f6f9; color: #1a1a2e; font-size: 14px; }}
  header {{ background: linear-gradient(135deg, #1a237e, #283593); color: white; padding: 24px 32px; }}
  header h1 {{ font-size: 22px; font-weight: 600; }}
  header p  {{ opacity: .8; margin-top: 4px; font-size: 13px; }}
  .container {{ padding: 24px 32px; max-width: 1400px; margin: auto; }}
  h2 {{ font-size: 15px; font-weight: 600; color: #283593; margin-bottom: 12px; text-transform: uppercase; letter-spacing: .5px; }}
  .cards {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 32px; }}
  .card {{ background: white; border-radius: 10px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,.08); border-top: 4px solid #9e9e9e; }}
  .card-falt {{ border-top-color: #d32f2f; }}
  .card-exc  {{ border-top-color: #f57c00; }}
  .card-ok   {{ border-top-color: #388e3c; }}
  .card-icon {{ font-size: 28px; margin-bottom: 8px; }}
  .card-title {{ font-size: 13px; font-weight: 600; color: #555; text-transform: uppercase; letter-spacing: .4px; margin-bottom: 12px; }}
  .card-nums {{ display: flex; gap: 20px; margin-bottom: 10px; }}
  .num-actual, .num-req {{ font-size: 28px; font-weight: 700; display: flex; flex-direction: column; }}
  .num-actual {{ color: #1a1a2e; }}
  .num-req    {{ color: #555; }}
  .num-actual small, .num-req small {{ font-size: 11px; font-weight: 400; color: #888; margin-top: 2px; }}
  .badge-falt {{ background: #ffebee; color: #c62828; padding: 4px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
  .badge-exc  {{ background: #fff3e0; color: #e65100; padding: 4px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
  .badge-ok   {{ background: #e8f5e9; color: #2e7d32; padding: 4px 10px; border-radius: 20px; font-size: 12px; font-weight: 600; }}
  .table-wrap {{ overflow-x: auto; background: white; border-radius: 10px; box-shadow: 0 2px 8px rgba(0,0,0,.08); }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
  thead th {{ background: #1a237e; color: white; padding: 10px 10px; text-align: center; font-weight: 500; white-space: nowrap; }}
  thead th.serv-header {{ text-align: left; }}
  tbody tr:nth-child(even) {{ background: #f8f9fc; }}
  tbody tr:hover {{ background: #e8eaf6; }}
  tbody tr.row-alert {{ background: #fff8f8; }}
  td {{ padding: 9px 10px; border-bottom: 1px solid #e8eaf6; text-align: center; }}
  td.serv-name {{ text-align: left; font-weight: 500; white-space: nowrap; }}
  td.obs {{ text-align: left; color: #555; font-size: 12px; min-width: 200px; }}
  td.falt {{ color: #c62828; font-weight: 700; background: #ffebee; }}
  td.exc  {{ color: #e65100; font-weight: 700; background: #fff3e0; }}
  td.ok   {{ color: #2e7d32; font-weight: 600; }}
  th.group {{ background: #283593; }}
  .criterios {{ background: white; border-radius: 10px; padding: 20px 24px; box-shadow: 0 2px 8px rgba(0,0,0,.08); margin-top: 24px; }}
  .criterios ul {{ columns: 2; gap: 32px; margin-top: 8px; }}
  .criterios li {{ margin-bottom: 4px; color: #444; list-style: none; padding-left: 16px; position: relative; }}
  .criterios li::before {{ content: '·'; position: absolute; left: 0; color: #1a237e; font-weight: 700; }}
  footer {{ text-align: center; padding: 16px; font-size: 12px; color: #888; margin-top: 24px; }}
  .row-sin-datos td {{ background: #fff8e1; color: #8d6e63; }}
  td.sin-datos {{ text-align: left; color: #e65100; background: #fff3e0; font-style: italic; font-size: 12px; font-weight: 600; }}
  .legend {{ display: flex; gap: 16px; margin-bottom: 16px; font-size: 12px; }}
  .legend span {{ padding: 3px 10px; border-radius: 12px; font-weight: 600; }}
</style>
</head>
<body>
<header>
  <h1>📊 Reporte de Infraestructura EDS — Estimación de necesidades</h1>
  <p>Archivo: {nombre_archivo} &nbsp;|&nbsp; Generado: {hoy} &nbsp;|&nbsp; Sistema: EDS IMSS-Bienestar</p>
</header>
<div class="container">

  <h2>Resumen global</h2>
  <div class="cards">{cards}</div>

  <h2>Desglose por servicio</h2>
  <div class="legend">
    <span style="background:#ffebee;color:#c62828">▲ Faltante</span>
    <span style="background:#fff3e0;color:#e65100">▼ Exceso sobre criterio</span>
    <span style="background:#e8f5e9;color:#2e7d32">✓ Cubierto</span>
  </div>
  <div class="table-wrap">
  <table>
    <thead>
      <tr>
        <th rowspan="2" class="serv-header">Servicio</th>
        <th colspan="4" class="group">🖥️ Computadoras</th>
        <th colspan="4" class="group">🖨️ Impresoras</th>
        <th colspan="4" class="group">📡 Access Points</th>
        <th colspan="4" class="group">🔌 Switches (24p)</th>
        <th rowspan="2">Observaciones</th>
      </tr>
      <tr>
        <th>Actual</th><th>Req.</th><th>Falta</th><th>Exceso</th>
        <th>Actual</th><th>Req.</th><th>Falta</th><th>Exceso</th>
        <th>Actual</th><th>Req.</th><th>Falta</th><th>Exceso</th>
        <th>Actual</th><th>Req.</th><th>Falta</th><th>Exceso</th>
      </tr>
    </thead>
    <tbody>{filas_serv}
    </tbody>
  </table>
  </div>

  <div class="criterios">
    <h2>Criterios de estimación aplicados</h2>
    <div style="background:#e8f5e9;border-left:4px solid #388e3c;padding:10px 16px;margin-bottom:14px;border-radius:4px;font-size:13px;color:#1b5e20;">
      <strong>Criterios oficiales IMSS-Bienestar</strong> (Criterios_Estimación_Equipos_de_Computo_e_Impresoras):<br>
      Computadoras e impresoras según tipo de área y número de camas/consultorios.<br><br>
      <strong>Criterios técnicos de red</strong> (buenas prácticas de la industria):<br>
      · <b>Access Point</b>: 1 por cada {COMPUTADORAS_POR_AP} dispositivos activos — rango recomendado por fabricantes es 20–30 clientes por radio para estabilidad en entornos empresariales.<br>
      · <b>Switch 24 puertos</b>: se reservan {PUERTOS_DE_RESERVA} puertos para uplink/interconexión, quedando {PUERTOS_POR_SWITCH - PUERTOS_DE_RESERVA} puertos útiles por switch — práctica estándar en redes LAN empresariales.
    </div>
    <ul>
      <li>Impresoras: 1 por cada {IMPRESORAS_POR_EQUIPOS} equipos de cómputo (criterio oficial)</li>
      <li>Access Point: 1 por cada {COMPUTADORAS_POR_AP} equipos (criterio técnico — rango recomendado: 20–30 dispositivos por AP)</li>
      <li>Switch 24p: 1 por cada {PUERTOS_POR_SWITCH - PUERTOS_DE_RESERVA} nodos ({PUERTOS_DE_RESERVA} puertos reservados para uplink)</li>
      <li>Urgencias — Módulo admisión/triage: 1 equipo x módulo</li>
      <li>Urgencias — Consultorios: 1 equipo x consultorio</li>
      <li>Urgencias — Camas (observación, choque): 1 equipo x 5 camas</li>
      <li>Urgencias — Central de enfermería: 1 equipo x central</li>
      <li>Hospitalización — Camas (hosp., UCI): 1 equipo x 5 camas</li>
      <li>Hospitalización — Central enfermería: 1 equipo x central</li>
      <li>Hospitalización — Farmacia hospitalaria: 1 equipo x área</li>
      <li>Quirófano — Salas: 1 equipo x 2 salas</li>
      <li>Quirófano — Central enfermería: 1 equipo x central</li>
      <li>Tococirugía — Salas expulsión: 1 equipo x 2 salas</li>
      <li>Tococirugía — Camas labor/recuperación: 1 equipo x 5 camas</li>
      <li>Consulta externa — Consultorios: 1 equipo x consultorio</li>
      <li>Consulta externa — Especialidades (Ginecología, Pediatría, etc.): 1 equipo x consultorio</li>
      <li>Consulta externa — Farmacia/ventanilla: 1 equipo x ventanilla</li>
      <li>Cuerpo de gobierno — Dirección, Subdirección, Jefaturas, Estadística, Epidemiología, Capturistas, Archivistas, Operadores: 1 equipo c/u (si hay personal)</li>
    </ul>
  </div>
</div>
<footer>Generado automáticamente por calcular_infraestructura.py — IMSS-Bienestar EDS</footer>
</body>
</html>"""

    if output_path:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html)
        print(f"\n  >> Reporte HTML generado: {output_path}")

    return html


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Calculadora de infraestructura EDS UM")
    parser.add_argument(
        "--archivo", type=str,
        default="Información implementación EDS UM.xlsx",
        help="Ruta al archivo Excel a procesar")
    parser.add_argument(
        "--actualizar-excel", action="store_true",
        help="Escribir AP/Switch faltantes en el bloque derecho de la hoja Red")
    parser.add_argument(
        "--generar-html", action="store_true",
        help="Generar reporte visual en HTML (abre en cualquier navegador)")

    # Override de totales (útil para inyectar valores del Excel resumen oficial)
    parser.add_argument("--comp-req", type=int, default=None,
                        help="Override: total de computadoras requeridas")
    parser.add_argument("--imp-req", type=int, default=None,
                        help="Override: total de impresoras requeridas")
    parser.add_argument("--ap-req", type=int, default=None,
                        help="Override: total de Access Points requeridos")
    parser.add_argument("--sw24-req", type=int, default=None,
                        help="Override: total de Switches 24p requeridos")

    args = parser.parse_args()

    overrides = {}
    if args.comp_req is not None:
        overrides["comp_req"] = args.comp_req
    if args.imp_req is not None:
        overrides["imp_req"] = args.imp_req
    if args.ap_req is not None:
        overrides["ap_req"] = args.ap_req
    if args.sw24_req is not None:
        overrides["sw24_req"] = args.sw24_req

    actual, requerido, faltantes, excesos, por_servicio, header_rows = procesar_archivo(args.archivo, overrides)
    generar_reporte(actual, requerido, faltantes, excesos, por_servicio)

    if args.actualizar_excel:
        print("\nActualizando Excel…")
        actualizar_excel_red(args.archivo, por_servicio, header_rows.get("Red", 4))

    if args.generar_html:
        print("\nGenerando reporte HTML…")
        base_nombre = os.path.splitext(os.path.basename(args.archivo))[0]
        out_path = os.path.join(os.getcwd(), base_nombre + "_reporte.html")
        generar_html(args.archivo, actual, requerido, faltantes, excesos, por_servicio,
                     output_path=out_path)


if __name__ == "__main__":
    main()
